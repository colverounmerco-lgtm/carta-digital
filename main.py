import os, csv, io, random, smtplib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date, time, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from functools import wraps
from pathlib import Path

EC_OFFSET = timedelta(hours=-5)   # Ecuador UTC-5, sin horario de verano

from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, make_response)
from sqlalchemy import text, inspect as sa_inspect, func
from dotenv import load_dotenv
import cloudinary, cloudinary.uploader

load_dotenv(Path(__file__).resolve().parent / ".env")
from werkzeug.security import generate_password_hash, check_password_hash
from models import db, Restaurante, Mesa, Producto, Orden, ItemOrden, MensajeSoporte, CodigoVerificacion, MetodoPago, Salsa, Adicion, SeccionBebida, VarianteBebida, SaborProducto, SubUsuario, ConfigGlobal, slugify

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "carta-dev-secret")

# ── Base de datos ──
_db_url = os.getenv("DATABASE_URL", "sqlite:///carta_local.db").strip()
if _db_url.startswith("postgres://"):
    _db_url = _db_url.replace("postgres://", "postgresql+pg8000://", 1)
elif _db_url.startswith("postgresql://") and "+pg8000" not in _db_url:
    _db_url = _db_url.replace("postgresql://", "postgresql+pg8000://", 1)

app.config["SQLALCHEMY_DATABASE_URI"] = _db_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

# ── Cloudinary ──
cloudinary.config(
    cloud_name=os.getenv("CLOUDINARY_CLOUD_NAME"),
    api_key=os.getenv("CLOUDINARY_API_KEY"),
    api_secret=os.getenv("CLOUDINARY_API_SECRET"),
)

MAX_CANTIDAD    = 10
TERMINOS_ASADO  = ["Blue", "Medio", "Tres cuartos", "Bien cocido"]
CODIGO_TTL   = 15  # minutos
TRIAL_DIAS   = 8   # Ecuador

TRIAL_DIAS_POR_PAIS = {"ecuador": 8, "colombia": 8}

PAGO_ECUADOR = {
    "whatsapp": "593968205068",
    "bancos": [
        {
            "banco":   "Banco Pichincha",
            "tipo":    "Ahorros",
            "cuenta":  "2215942646",
            "titular": "Santiago Medina",
            "cedula":  "1753446341",
        },
        {
            "banco":   "Banco Guayaquil",
            "tipo":    "Ahorros",
            "cuenta":  "0038528429",
            "titular": "Santiago Medina",
            "cedula":  "1753446341",
        },
    ],
}

PLANES = {
    "mensual": {"nombre": "Mensual", "dias": 30},
    "anual":   {"nombre": "Anual",   "dias": 365},
}

PRECIOS = {
    "ecuador": {
        "mensual": {"total": 33.35,    "simbolo": "$",   "moneda": "USD", "desc": "$33.35/mes · IVA incluido",         "desc_mes": "$33.35/mes"},
        "anual":   {"total": 333.50,   "simbolo": "$",   "moneda": "USD", "desc": "$333.50/año · IVA incluido",        "desc_mes": "$27.79/mes"},
    },
    "colombia": {
        "mensual": {"total": 117810,   "simbolo": "COP", "moneda": "COP", "desc": "COP 117.810/mes · IVA incluido",    "desc_mes": "COP 117.810/mes"},
        "anual":   {"total": 1178100,  "simbolo": "COP", "moneda": "COP", "desc": "COP 1.178.100/año · IVA incluido", "desc_mes": "COP 98.175/mes"},
    },
}


def plan_vigente(r):
    """True si el restaurante tiene acceso activo."""
    if not r:
        return False
    if not r.plan_vence:
        return True
    return datetime.utcnow() <= r.plan_vence


def dias_plan(r):
    """Días restantes del plan (None si no hay fecha de vencimiento)."""
    if not r or not r.plan_vence:
        return None
    delta = r.plan_vence - datetime.utcnow()
    return max(0, delta.days)


def _crear_metodos_default(restaurante_id):
    defaults = [("💵", "Efectivo", 0), ("💳", "Tarjeta", 1), ("📲", "Transferencia", 2), ("/static/img/deuna.svg", "Deuna!", 3)]
    for icono, nombre, orden in defaults:
        db.session.add(MetodoPago(
            restaurante_id=restaurante_id, nombre=nombre,
            icono=icono, orden_display=orden
        ))


def inicio_fin_dia_ec():
    """Devuelve (inicio, fin) del día actual en Ecuador como datetimes UTC."""
    ec_hoy = (datetime.utcnow() + EC_OFFSET).date()
    inicio = datetime.combine(ec_hoy, time.min) - EC_OFFSET
    fin    = datetime.combine(ec_hoy, time.max) - EC_OFFSET
    return inicio, fin


def get_client_ip():
    """Devuelve la IP real del cliente, pasando por proxies de Railway/nginx."""
    xff = request.headers.get("X-Forwarded-For", "")
    if xff:
        return xff.split(",")[0].strip()
    return request.remote_addr


# ── Email ──
def enviar_email(destinatario, asunto, cuerpo_html):
    # Prioridad 1: Resend API (funciona en Railway y cualquier cloud)
    resend_key = os.getenv("RESEND_API_KEY", "")
    if resend_key:
        import urllib.request, json as _json
        payload = _json.dumps({
            "from":    os.getenv("RESEND_FROM", "Carta Digital <noreply@cartadigital.app>"),
            "to":      [destinatario],
            "subject": asunto,
            "html":    cuerpo_html,
        }).encode("utf-8")
        req = urllib.request.Request(
            "https://api.resend.com/emails",
            data=payload,
            headers={
                "Authorization": f"Bearer {resend_key}",
                "Content-Type":  "application/json",
            },
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            print(f"[EMAIL OK → {destinatario}] {asunto} (Resend {resp.status})")
        return

    # Prioridad 2: SMTP directo (funciona en local)
    host = os.getenv("SMTP_HOST", "")
    user = os.getenv("SMTP_USER", "")
    pwd  = os.getenv("SMTP_PASS", "")
    port = int(os.getenv("SMTP_PORT", "587"))

    if not all([host, user, pwd]):
        print(f"\n[EMAIL → {destinatario}] {asunto}\n{cuerpo_html}\n")
        return

    msg = MIMEMultipart("alternative")
    msg["Subject"] = asunto
    msg["From"]    = f"Carta Digital <{user}>"
    msg["To"]      = destinatario
    msg.attach(MIMEText(cuerpo_html, "html", "utf-8"))

    with smtplib.SMTP(host, port, timeout=15) as srv:
        srv.ehlo()
        srv.starttls()
        srv.login(user, pwd)
        srv.sendmail(user, destinatario, msg.as_string())
        print(f"[EMAIL OK → {destinatario}] {asunto}")


def generar_codigo():
    return str(random.randint(100000, 999999))


def crear_codigo(email, tipo):
    # Invalida códigos anteriores del mismo tipo
    CodigoVerificacion.query.filter_by(email=email, tipo=tipo, usado=False).update({"usado": True})
    codigo = generar_codigo()
    db.session.add(CodigoVerificacion(
        email=email, codigo=codigo, tipo=tipo,
        expira=datetime.utcnow() + timedelta(minutes=CODIGO_TTL)
    ))
    db.session.commit()
    return codigo


def validar_codigo(email, codigo, tipo):
    c = CodigoVerificacion.query.filter_by(
        email=email, codigo=codigo, tipo=tipo, usado=False
    ).first()
    if not c or datetime.utcnow() > c.expira:
        return False
    c.usado = True
    db.session.commit()
    return True


# ── Helpers ──
def restaurante_session():
    rid = session.get("restaurante_id")
    return Restaurante.query.get(rid) if rid else None


def subir_imagen(file_storage, folder="carta_digital"):
    # Fallback local cuando Cloudinary no está configurado
    if os.getenv("CLOUDINARY_CLOUD_NAME", "") in ("", "placeholder"):
        import uuid
        from werkzeug.utils import secure_filename
        ext      = Path(file_storage.filename).suffix.lower() or ".jpg"
        filename = f"{uuid.uuid4().hex}{ext}"
        upload_dir = Path(__file__).parent / "static" / "uploads"
        upload_dir.mkdir(exist_ok=True)
        file_storage.save(upload_dir / filename)
        return f"/static/uploads/{filename}"

    result = cloudinary.uploader.upload(
        file_storage,
        folder=folder,
        transformation=[{"width": 900, "height": 700, "crop": "fill", "quality": "auto:good"}],
    )
    return result["secure_url"]


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("restaurante_id"):
            flash("Inicia sesión primero.", "error")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def plan_requerido(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        r = restaurante_session()
        if r and not r.activo:
            session.clear()
            flash("Tu cuenta está desactivada. Contacta al administrador.", "error")
            return redirect(url_for("login"))
        if r and not plan_vigente(r):
            vencido_hace = (datetime.utcnow() - r.plan_vence).days if r.plan_vence else 0
            return render_template("auth/plan_vencido.html",
                                   restaurante=r, vencido_hace=vencido_hace,
                                   precios=PRECIOS.get(r.pais or 'ecuador', PRECIOS['ecuador']),
                                   pago_ecuador=PAGO_ECUADOR)
        return f(*args, **kwargs)
    return decorated


@app.context_processor
def ctx():
    sid = session.get("subusuario_id")
    subusuario = SubUsuario.query.get(sid) if sid else None

    def fp(val, restaurante=None):
        pais = getattr(restaurante, 'pais', None) if restaurante else None
        if pais == 'colombia':
            return "COP " + "{:,.0f}".format(val).replace(",", ".")
        return "${:.2f}".format(val)

    return {
        "restaurante_session": restaurante_session(),
        "subusuario_actual": subusuario,
        "now": datetime.utcnow,
        "fp": fp,
    }


# Endpoints accesibles por subusuarios (los demás requieren ser dueño)
_ENDPOINTS_STAFF = {
    "dashboard", "confirmar_orden", "orden_lista", "pagar_orden",
    "cancelar_orden", "cerrar_cuenta_mesa", "api_ordenes_activas", "logout", "staff_logout",
    "static", "recibo_orden", "staff_login", "staff_login_slug",
}

@app.before_request
def verificar_acceso_subusuario():
    if not session.get("subusuario_id"):
        return
    endpoint = request.endpoint or ""
    if endpoint not in _ENDPOINTS_STAFF:
        flash("No tienes acceso a esa sección.", "error")
        return redirect(url_for("dashboard"))


@app.template_filter("ec_time")
def ec_time_filter(dt, fmt="%d/%m/%Y %H:%M"):
    if not dt:
        return ""
    return (dt + EC_OFFSET).strftime(fmt)


# ── Migraciones runtime ──
def run_migrations():
    insp   = sa_inspect(db.engine)
    tables = insp.get_table_names()

    # Cache de columnas por tabla para no hacer una query por cada add_col
    _col_cache = {}
    def cols_of(table):
        if table not in _col_cache:
            _col_cache[table] = {c["name"] for c in insp.get_columns(table)}
        return _col_cache[table]

    def add_col(table, col, definition):
        if col not in cols_of(table):
            db.session.execute(text(f"ALTER TABLE {table} ADD COLUMN {col} {definition}"))
            db.session.commit()
            _col_cache[table].add(col)  # actualizar cache local

    # Solo ejecutar ALTER TYPE si la columna aún es VARCHAR corta
    if "metodos_pago" in tables:
        icono_col = next((c for c in insp.get_columns("metodos_pago") if c["name"] == "icono"), None)
        col_length = getattr(icono_col["type"], "length", None) if icono_col else None
        if col_length is not None and col_length < 300:
            db.session.execute(text("ALTER TABLE metodos_pago ALTER COLUMN icono TYPE VARCHAR(300)"))
            db.session.commit()

    if "ordenes" in tables:
        add_col("ordenes", "metodo_pago",      "VARCHAR(30)")
        add_col("ordenes", "notas",            "TEXT")
        add_col("ordenes", "solicita_cuenta",  "BOOLEAN DEFAULT FALSE")
        add_col("ordenes", "metodo_preferido", "VARCHAR(30)")
        add_col("ordenes", "fecha_pago",       "TIMESTAMP")

    if "mesas" in tables:
        nuevas_mesas = [
            ("abierta",           "BOOLEAN DEFAULT FALSE"),
            ("es_para_llevar",    "BOOLEAN DEFAULT FALSE"),
            ("tab_inicio",        "TIMESTAMP"),
            ("mesero_solicitado", "BOOLEAN DEFAULT FALSE"),
        ]
        agregadas = any(col not in cols_of("mesas") for col, _ in nuevas_mesas)
        for col, defn in nuevas_mesas:
            add_col("mesas", col, defn)
        # Solo limpiar mesas en el primer despliegue donde se añadió la columna "abierta"
        if agregadas:
            db.session.execute(text(
                "UPDATE mesas SET abierta = FALSE "
                "WHERE id NOT IN ("
                "  SELECT DISTINCT mesa_id FROM ordenes "
                "  WHERE estado IN ('pendiente','confirmada','lista')"
                ")"
            ))
            db.session.commit()

    if "productos" in tables:
        add_col("productos", "orden_display",     "INTEGER DEFAULT 0")
        add_col("productos", "terminos_asado",    "BOOLEAN DEFAULT FALSE")
        add_col("productos", "salsas_activas",    "BOOLEAN DEFAULT FALSE")
        add_col("productos", "adiciones_activas", "BOOLEAN DEFAULT FALSE")
        add_col("productos", "bebidas_activas",   "BOOLEAN DEFAULT FALSE")
        add_col("productos", "sabores_activos",   "BOOLEAN DEFAULT FALSE")

    if "restaurantes" in tables:
        add_col("restaurantes", "descripcion",       "TEXT")
        add_col("restaurantes", "email_verificado",  "BOOLEAN DEFAULT FALSE")
        add_col("restaurantes", "plan",              "VARCHAR(20) DEFAULT 'trial'")
        add_col("restaurantes", "plan_inicio",       "TIMESTAMP")
        add_col("restaurantes", "plan_vence",        "TIMESTAMP")
        add_col("restaurantes", "ip_red",            "VARCHAR(50)")
        add_col("restaurantes", "restringir_red",    "BOOLEAN DEFAULT TRUE")
        add_col("restaurantes", "dia_apertura",      "DATE")
        add_col("restaurantes", "modo_cobro",        "BOOLEAN DEFAULT FALSE")
        add_col("restaurantes", "categoria",         "VARCHAR(20) DEFAULT 'restaurante'")
        add_col("restaurantes", "pais",              "VARCHAR(20) DEFAULT 'ecuador'")
        add_col("restaurantes", "fact_tipo_id",      "VARCHAR(20)")
        add_col("restaurantes", "fact_numero_id",    "VARCHAR(30)")
        add_col("restaurantes", "fact_razon_social", "VARCHAR(150)")
        add_col("restaurantes", "fact_direccion",    "VARCHAR(200)")

    # ── Datos por defecto: una sola query por tabla ──
    # IDs de restaurantes que ya tienen Deuna!
    ids_con_deuna = {
        row[0] for row in
        db.session.execute(text("SELECT restaurante_id FROM metodos_pago WHERE nombre='Deuna!'")).fetchall()
    }
    # IDs de restaurantes que no tienen ningún método de pago
    ids_sin_metodos = {
        row[0] for row in
        db.session.execute(text(
            "SELECT r.id FROM restaurantes r "
            "LEFT JOIN metodos_pago mp ON mp.restaurante_id = r.id "
            "GROUP BY r.id HAVING COUNT(mp.id) = 0"
        )).fetchall()
    }

    restaurantes = Restaurante.query.all()
    for r in restaurantes:
        if r.id in ids_sin_metodos:
            _crear_metodos_default(r.id)
        if r.id not in ids_con_deuna:
            ultimo = MetodoPago.query.filter_by(restaurante_id=r.id).count()
            db.session.add(MetodoPago(restaurante_id=r.id, nombre="Deuna!", icono="/static/img/deuna.svg", orden_display=ultimo))
        if not r.plan:
            r.plan = 'trial'
        dias_correctos = TRIAL_DIAS_POR_PAIS.get(r.pais or 'ecuador', TRIAL_DIAS)
        if not r.plan_vence:
            r.plan_vence = (r.fecha_registro or datetime.utcnow()) + timedelta(days=dias_correctos)
        elif r.plan == 'trial' and r.plan_inicio:
            if (r.plan_vence - r.plan_inicio).days != dias_correctos:
                r.plan_vence = r.plan_inicio + timedelta(days=dias_correctos)
    db.session.commit()


with app.app_context():
    db.create_all()
    run_migrations()
    # Auto-create "Para llevar" mesa for every restaurant that doesn't have one
    for _r in Restaurante.query.all():
        if not Mesa.query.filter_by(restaurante_id=_r.id, es_para_llevar=True).first():
            db.session.add(Mesa(
                restaurante_id=_r.id,
                numero=0,
                nombre="Para llevar",
                token=Mesa.nuevo_token(),
                activa=True,
                abierta=True,
                es_para_llevar=True,
            ))
    db.session.commit()


# ══════════════════════════════════════════════
#  AUTH
# ══════════════════════════════════════════════

@app.route("/")
def index():
    if session.get("restaurante_id"):
        return redirect(url_for("dashboard"))
    return render_template("landing.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        nombre      = request.form.get("nombre", "").strip()
        email       = request.form.get("email", "").strip().lower()
        pwd         = request.form.get("password", "")
        pwd2        = request.form.get("password2", "")
        whatsapp    = request.form.get("whatsapp", "").strip()
        ciudad      = request.form.get("ciudad", "").strip()
        descripcion = request.form.get("descripcion", "").strip()
        categoria         = request.form.get("categoria", "").strip()
        pais              = request.form.get("pais", "").strip()
        fact_tipo_id      = request.form.get("fact_tipo_id", "").strip()
        fact_numero_id    = request.form.get("fact_numero_id", "").strip()
        fact_razon_social = request.form.get("fact_razon_social", "").strip()
        fact_direccion    = request.form.get("fact_direccion", "").strip()

        if not all([nombre, email, pwd, categoria, pais]):
            flash("Completa todos los campos requeridos.", "error")
            return redirect(url_for("register"))
        if categoria not in ("restaurante", "bar", "cafeteria"):
            flash("Selecciona un tipo de local válido.", "error")
            return redirect(url_for("register"))
        if pais not in ("ecuador", "colombia"):
            flash("Selecciona un país válido.", "error")
            return redirect(url_for("register"))
        if pwd != pwd2:
            flash("Las contraseñas no coinciden.", "error")
            return redirect(url_for("register"))
        if len(pwd) < 6:
            flash("La contraseña debe tener al menos 6 caracteres.", "error")
            return redirect(url_for("register"))
        if Restaurante.query.filter_by(email=email).first():
            flash("Ya existe una cuenta con ese correo.", "error")
            return redirect(url_for("register"))

        base_slug, counter, slug = slugify(nombre), 1, slugify(nombre)
        while Restaurante.query.filter_by(slug=slug).first():
            slug = f"{base_slug}-{counter}"; counter += 1

        logo_url = None
        if "logo" in request.files and request.files["logo"].filename:
            try:
                logo_url = subir_imagen(request.files["logo"], "carta_logos")
            except Exception as e:
                flash(f"Error al subir logo: {e}", "error")
                return redirect(url_for("register"))

        skip_verificacion = os.getenv("SKIP_EMAIL_VERIFICATION", "false").lower() == "true"

        trial_dias = TRIAL_DIAS_POR_PAIS.get(pais, TRIAL_DIAS)
        r = Restaurante(nombre=nombre, email=email, slug=slug,
                        whatsapp=whatsapp, ciudad=ciudad,
                        logo_url=logo_url, descripcion=descripcion,
                        categoria=categoria, pais=pais,
                        plan='trial',
                        plan_inicio=datetime.utcnow(),
                        plan_vence=datetime.utcnow() + timedelta(days=trial_dias),
                        email_verificado=skip_verificacion,
                        fact_tipo_id=fact_tipo_id or None,
                        fact_numero_id=fact_numero_id or None,
                        fact_razon_social=fact_razon_social or None,
                        fact_direccion=fact_direccion or None)
        r.set_password(pwd)
        db.session.add(r)
        db.session.flush()
        _crear_metodos_default(r.id)
        db.session.add(Mesa(
            restaurante_id=r.id,
            numero=0,
            nombre="Para llevar",
            token=Mesa.nuevo_token(),
            activa=True,
            abierta=True,
            es_para_llevar=True,
        ))
        db.session.commit()

        if skip_verificacion:
            session["restaurante_id"] = r.id
            flash(f"¡Bienvenido, {r.nombre}! Empieza creando tu menú.", "success")
            return redirect(url_for("dashboard"))

        codigo = crear_codigo(email, "registro")
        try:
            enviar_email(email, "Verifica tu cuenta — Carta Digital",
                f"""<div style="font-family:Inter,sans-serif;max-width:480px;margin:0 auto;padding:2rem">
                <h2 style="color:#F97316">¡Bienvenido a Carta Digital!</h2>
                <p>Usa este código para verificar tu cuenta:</p>
                <div style="font-size:2.5rem;font-weight:900;letter-spacing:0.5rem;
                            text-align:center;padding:1.5rem;background:#F9FAFB;
                            border-radius:12px;margin:1.5rem 0;color:#1C1C1E">
                  {codigo}
                </div>
                <p style="color:#6B7280;font-size:0.85rem">
                  Este código expira en {CODIGO_TTL} minutos.<br>
                  Si no creaste esta cuenta, ignora este correo.
                </p></div>""")
        except Exception as e:
            print(f"[EMAIL ERROR] {e}")

        session["verificar_email"] = email
        flash("Te enviamos un código de verificación a tu correo.", "info")
        return redirect(url_for("verificar_email"))

    return render_template("auth/register.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        pwd   = request.form.get("password", "")
        r = Restaurante.query.filter_by(email=email).first()
        if not r or not r.check_password(pwd):
            flash("Correo o contraseña incorrectos.", "error")
            return redirect(url_for("login"))
        if not r.activo:
            flash("Tu cuenta está desactivada.", "error")
            return redirect(url_for("login"))
        if not r.email_verificado:
            codigo = crear_codigo(email, "registro")
            try:
                enviar_email(email, "Verifica tu cuenta — Carta Digital",
                    f"""<div style="font-family:Inter,sans-serif;max-width:480px;margin:0 auto;padding:2rem">
                    <h2 style="color:#F97316">Verifica tu cuenta</h2>
                    <p>Tu cuenta aún no está verificada. Usa este código:</p>
                    <div style="font-size:2.5rem;font-weight:900;letter-spacing:0.5rem;
                                text-align:center;padding:1.5rem;background:#F9FAFB;
                                border-radius:12px;margin:1.5rem 0;color:#1C1C1E">
                      {codigo}
                    </div>
                    <p style="color:#6B7280;font-size:0.85rem">Expira en {CODIGO_TTL} minutos.</p>
                    </div>""")
            except Exception as e:
                print(f"[EMAIL ERROR] {e}")
            session["verificar_email"] = email
            flash("Verifica tu correo antes de entrar.", "error")
            return redirect(url_for("verificar_email"))
        session["restaurante_id"] = r.id
        return redirect(url_for("dashboard"))
    return render_template("auth/login.html")


@app.route("/logout")
def logout():
    session.pop("restaurante_id",  None)
    session.pop("subusuario_id",   None)
    session.pop("subusuario_rol",  None)
    return redirect(url_for("index"))


# ── Login de staff (cocineros / meseros) ──

@app.route("/staff")
def staff_login():
    return redirect(url_for("login"))  # fallback: sin slug, redirige al login normal


@app.route("/staff/<slug>", methods=["GET", "POST"])
def staff_login_slug(slug):
    if session.get("subusuario_id"):
        return redirect(url_for("dashboard"))
    r = Restaurante.query.filter_by(slug=slug, activo=True).first_or_404()
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        pwd      = request.form.get("password", "")
        su = SubUsuario.query.filter_by(restaurante_id=r.id, username=username, activo=True).first()
        if su and su.check_password(pwd):
            session["restaurante_id"] = r.id
            session["subusuario_id"]  = su.id
            session["subusuario_rol"] = su.rol
            return redirect(url_for("dashboard"))
        error = "Usuario o contraseña incorrectos."
    return render_template("auth/staff_login.html", restaurante=r, error=error)


@app.route("/staff/logout", methods=["POST"])
def staff_logout():
    session.pop("restaurante_id", None)
    session.pop("subusuario_id",  None)
    session.pop("subusuario_rol", None)
    return redirect(url_for("staff_login"))


# ── Gestión de subusuarios (solo dueño) ──

@app.route("/subusuarios")
@login_required
@plan_requerido
def subusuarios():
    r = restaurante_session()
    lista = SubUsuario.query.filter_by(restaurante_id=r.id).order_by(SubUsuario.rol, SubUsuario.nombre).all()
    return render_template("restaurante/subusuarios.html", restaurante=r, subusuarios=lista)


@app.route("/subusuarios/agregar", methods=["POST"])
@login_required
def agregar_subusuario():
    r        = restaurante_session()
    nombre   = request.form.get("nombre", "").strip()
    username = request.form.get("username", "").strip().lower()
    pwd      = request.form.get("password", "")
    rol      = request.form.get("rol", "cocinero")
    if not all([nombre, username, pwd]):
        flash("Completa todos los campos.", "error")
        return redirect(url_for("subusuarios"))
    if rol not in ("cocinero", "mesero"):
        flash("Rol inválido.", "error")
        return redirect(url_for("subusuarios"))
    if SubUsuario.query.filter_by(restaurante_id=r.id, username=username).first():
        flash(f"El usuario '{username}' ya existe en tu restaurante.", "error")
        return redirect(url_for("subusuarios"))
    su = SubUsuario(restaurante_id=r.id, nombre=nombre, username=username, rol=rol)
    su.set_password(pwd)
    db.session.add(su)
    db.session.commit()
    flash(f"{nombre} agregado como {rol}.", "success")
    return redirect(url_for("subusuarios"))


@app.route("/subusuarios/<int:sid>/toggle", methods=["POST"])
@login_required
def toggle_subusuario(sid):
    r  = restaurante_session()
    su = SubUsuario.query.filter_by(id=sid, restaurante_id=r.id).first_or_404()
    su.activo = not su.activo
    db.session.commit()
    estado = "activado" if su.activo else "desactivado"
    flash(f"{su.nombre} {estado}.", "success")
    return redirect(url_for("subusuarios"))


@app.route("/subusuarios/<int:sid>/eliminar", methods=["POST"])
@login_required
def eliminar_subusuario(sid):
    r  = restaurante_session()
    su = SubUsuario.query.filter_by(id=sid, restaurante_id=r.id).first_or_404()
    db.session.delete(su)
    db.session.commit()
    flash(f"{su.nombre} eliminado.", "success")
    return redirect(url_for("subusuarios"))


@app.route("/subusuarios/<int:sid>/cambiar-password", methods=["POST"])
@login_required
def subusuario_cambiar_password(sid):
    r          = restaurante_session()
    su         = SubUsuario.query.filter_by(id=sid, restaurante_id=r.id).first_or_404()
    pwd_actual = request.form.get("password_actual", "")
    pwd        = request.form.get("password", "")
    pwd2       = request.form.get("password2", "")
    if not r.check_password(pwd_actual):
        flash("Tu contraseña actual es incorrecta.", "error")
    elif pwd != pwd2:
        flash("Las contraseñas no coinciden.", "error")
    elif len(pwd) < 6:
        flash("La contraseña debe tener al menos 6 caracteres.", "error")
    else:
        su.set_password(pwd)
        db.session.commit()
        flash(f"Contraseña de {su.nombre} actualizada.", "success")
    return redirect(url_for("subusuarios"))


@app.route("/verificar-email", methods=["GET", "POST"])
def verificar_email():
    email = session.get("verificar_email")
    if not email:
        return redirect(url_for("login"))

    if request.method == "POST":
        codigo = request.form.get("codigo", "").strip()
        if validar_codigo(email, codigo, "registro"):
            r = Restaurante.query.filter_by(email=email).first()
            if r:
                r.email_verificado = True
                db.session.commit()
                session.pop("verificar_email", None)
                session["restaurante_id"] = r.id
                flash(f"¡Bienvenido, {r.nombre}! Empieza creando tu menú.", "success")
                return redirect(url_for("dashboard"))
        flash("Código incorrecto o expirado.", "error")
        return redirect(url_for("verificar_email"))

    return render_template("auth/verificar_email.html", email=email)


@app.route("/verificar-email/reenviar", methods=["POST"])
def reenviar_codigo_registro():
    email = session.get("verificar_email")
    if not email:
        return redirect(url_for("login"))
    codigo = crear_codigo(email, "registro")
    try:
        enviar_email(email, "Tu nuevo código — Carta Digital",
            f"""<div style="font-family:Inter,sans-serif;max-width:480px;margin:0 auto;padding:2rem">
            <h2 style="color:#F97316">Nuevo código de verificación</h2>
            <div style="font-size:2.5rem;font-weight:900;letter-spacing:0.5rem;
                        text-align:center;padding:1.5rem;background:#F9FAFB;
                        border-radius:12px;margin:1.5rem 0;color:#1C1C1E">
              {codigo}
            </div>
            <p style="color:#6B7280;font-size:0.85rem">Expira en {CODIGO_TTL} minutos.</p>
            </div>""")
        flash("Te enviamos un nuevo código.", "success")
    except Exception as e:
        print(f"[EMAIL ERROR] {e}")
        flash("Error al enviar el correo. Intenta de nuevo.", "error")
    return redirect(url_for("verificar_email"))


@app.route("/recuperar", methods=["GET", "POST"])
def recuperar():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        r = Restaurante.query.filter_by(email=email).first()
        if r:
            codigo = crear_codigo(email, "reset")
            try:
                enviar_email(email, "Recupera tu contraseña — Carta Digital",
                    f"""<div style="font-family:Inter,sans-serif;max-width:480px;margin:0 auto;padding:2rem">
                    <h2 style="color:#F97316">Recuperar contraseña</h2>
                    <p>Usa este código para crear una nueva contraseña:</p>
                    <div style="font-size:2.5rem;font-weight:900;letter-spacing:0.5rem;
                                text-align:center;padding:1.5rem;background:#F9FAFB;
                                border-radius:12px;margin:1.5rem 0;color:#1C1C1E">
                      {codigo}
                    </div>
                    <p style="color:#6B7280;font-size:0.85rem">
                      Expira en {CODIGO_TTL} minutos.<br>
                      Si no solicitaste esto, ignora este correo.
                    </p></div>""")
            except Exception as e:
                print(f"[EMAIL ERROR] {e}")
        # Siempre mostramos el mismo mensaje para no revelar si el email existe
        session["reset_email"] = email
        flash("Si ese correo está registrado, recibirás el código en breve.", "info")
        return redirect(url_for("recuperar_codigo"))
    return render_template("auth/recuperar.html")


@app.route("/recuperar/codigo", methods=["GET", "POST"])
def recuperar_codigo():
    email = session.get("reset_email")
    if not email:
        return redirect(url_for("recuperar"))

    if request.method == "POST":
        codigo = request.form.get("codigo", "").strip()
        pwd    = request.form.get("password", "")
        pwd2   = request.form.get("password2", "")

        if not codigo:
            flash("Ingresa el código.", "error")
            return redirect(url_for("recuperar_codigo"))
        if pwd != pwd2:
            flash("Las contraseñas no coinciden.", "error")
            return redirect(url_for("recuperar_codigo"))
        if len(pwd) < 6:
            flash("La contraseña debe tener al menos 6 caracteres.", "error")
            return redirect(url_for("recuperar_codigo"))
        if not validar_codigo(email, codigo, "reset"):
            flash("Código incorrecto o expirado.", "error")
            return redirect(url_for("recuperar_codigo"))

        r = Restaurante.query.filter_by(email=email).first()
        if r:
            r.set_password(pwd)
            db.session.commit()
        session.pop("reset_email", None)
        flash("Contraseña actualizada. Ya puedes iniciar sesión.", "success")
        return redirect(url_for("login"))

    return render_template("auth/recuperar_codigo.html", email=email)


# ══════════════════════════════════════════════
#  DASHBOARD
# ══════════════════════════════════════════════

@app.route("/dashboard")
@login_required
@plan_requerido
def dashboard():
    r = restaurante_session()
    es_sub = bool(session.get("subusuario_id"))

    if not es_sub:
        # Solo el dueño actualiza IP y abre mesas al inicio del día
        ip_actual = get_client_ip()
        if r.ip_red != ip_actual:
            r.ip_red = ip_actual
        hoy_ec = (datetime.utcnow() + EC_OFFSET).date()
        if r.dia_apertura != hoy_ec:
            Mesa.query.filter_by(restaurante_id=r.id, activa=True).update({"abierta": True})
            r.dia_apertura = hoy_ec
        db.session.commit()
    inicio, fin     = inicio_fin_dia_ec()

    estados_activos = ["pendiente", "confirmada"] if r.categoria == 'bar' else ["pendiente", "confirmada", "lista"]
    ordenes_activas = Orden.query.filter(
        Orden.restaurante_id == r.id,
        Orden.estado.in_(estados_activos),
        Orden.fecha >= inicio,
        Orden.fecha <= fin,
    ).order_by(Orden.fecha.desc()).all()

    total_hoy = db.session.query(func.count(Orden.id)).filter(
        Orden.restaurante_id == r.id,
        Orden.estado == "pagada",
        Orden.fecha >= inicio,
        Orden.fecha <= fin,
    ).scalar() or 0

    ingresos_hoy = db.session.query(func.sum(Orden.total)).filter(
        Orden.restaurante_id == r.id,
        Orden.estado == "pagada",
        Orden.fecha >= inicio,
        Orden.fecha <= fin,
    ).scalar() or 0.0

    cuentas_solicitadas  = sum(1 for o in ordenes_activas if o.solicita_cuenta)
    meseros_solicitados  = Mesa.query.filter_by(restaurante_id=r.id, mesero_solicitado=True).all()
    metodos_pago = MetodoPago.query.filter_by(
        restaurante_id=r.id, activo=True
    ).order_by(MetodoPago.orden_display).all()

    # Para bares: agrupar TODOS los pedidos de hoy por mesa (activos + entregados)
    # Solo mostramos mesas que tengan al menos un pedido no cerrado (pendiente/confirmada/lista)
    cuentas_abiertas = []
    if r.categoria == 'bar':
        ordenes_bar_abiertas = Orden.query.filter(
            Orden.restaurante_id == r.id,
            Orden.estado.in_(["pendiente", "confirmada", "lista"]),
            Orden.fecha >= inicio,
            Orden.fecha <= fin,
        ).all()
        mesas_con_activos = {o.mesa_id for o in ordenes_bar_abiertas}
        grupos = {}
        for mid in mesas_con_activos:
            mesa       = Mesa.query.get(mid)
            desde      = mesa.tab_inicio if mesa and mesa.tab_inicio else inicio
            todas = Orden.query.filter(
                Orden.mesa_id == mid,
                Orden.restaurante_id == r.id,
                Orden.estado != 'cancelada',
                Orden.fecha >= desde,
            ).order_by(Orden.fecha).all()
            grupos[mid] = {
                "mesa":    mesa,
                "ordenes": todas,
                "total":   sum(o.total for o in todas),
            }
        cuentas_abiertas = list(grupos.values())

    return render_template("restaurante/dashboard.html",
        restaurante=r,
        ordenes_activas=ordenes_activas,
        total_hoy=total_hoy,
        ingresos_hoy=ingresos_hoy,
        cuentas_solicitadas=cuentas_solicitadas,
        meseros_solicitados=meseros_solicitados,
        metodos_pago=metodos_pago,
        dias_plan=dias_plan(r),
        es_subusuario=es_sub,
        rol_actual=session.get("subusuario_rol"),
        cuentas_abiertas=cuentas_abiertas,
    )


# ── Acciones sobre órdenes ──

@app.route("/dashboard/orden/<int:oid>/confirmar", methods=["POST"])
@login_required
def confirmar_orden(oid):
    rol = session.get("subusuario_rol")
    if rol == "mesero":
        return redirect(url_for("dashboard"))
    r = restaurante_session()
    # En cobro anticipado solo el dueño confirma (= cobró en caja)
    if rol == "cocinero" and r.modo_cobro:
        return redirect(url_for("dashboard"))
    o = Orden.query.filter_by(id=oid, restaurante_id=r.id).first_or_404()
    if o.estado == "pendiente":
        o.estado = "confirmada"
        db.session.commit()
    return redirect(url_for("dashboard"))


@app.route("/dashboard/orden/<int:oid>/lista", methods=["POST"])
@login_required
def orden_lista(oid):
    rol = session.get("subusuario_rol")
    if rol == "mesero":
        return redirect(url_for("dashboard"))
    r = restaurante_session()
    o = Orden.query.filter_by(id=oid, restaurante_id=r.id).first_or_404()
    if o.estado == "confirmada":
        o.estado = "lista"
        db.session.commit()
    return redirect(url_for("dashboard"))


@app.route("/dashboard/orden/<int:oid>/pagar", methods=["POST"])
@login_required
def pagar_orden(oid):
    rol = session.get("subusuario_rol")
    if rol == "cocinero":
        return redirect(url_for("dashboard"))
    r = restaurante_session()
    o = Orden.query.filter_by(id=oid, restaurante_id=r.id).first_or_404()
    if o.estado == "lista":
        o.estado      = "pagada"
        o.metodo_pago = request.form.get("metodo", "efectivo")
        o.fecha_pago  = datetime.utcnow()
        mesa = Mesa.query.get(o.mesa_id)
        if mesa and r.categoria != 'bar':
            mesa.abierta = False
        db.session.commit()
    return redirect(url_for("dashboard"))


@app.route("/dashboard/mesa/<int:mid>/cerrar-cuenta", methods=["POST"])
@login_required
def cerrar_cuenta_mesa(mid):
    r    = restaurante_session()
    mesa = Mesa.query.filter_by(id=mid, restaurante_id=r.id).first_or_404()
    metodo = request.form.get("metodo", "efectivo")
    ordenes = Orden.query.filter(
        Orden.mesa_id == mid,
        Orden.estado.in_(["pendiente", "confirmada", "lista"]),
    ).all()
    for o in ordenes:
        o.estado      = "pagada"
        o.metodo_pago = metodo
        o.fecha_pago  = datetime.utcnow()
    mesa.abierta    = False
    mesa.tab_inicio = None
    db.session.commit()
    flash(f"Cuenta de {mesa.nombre} cerrada.", "success")
    return redirect(url_for("dashboard"))


@app.route("/dashboard/mesa/<int:mid>/atiende-mesero", methods=["POST"])
@login_required
def atiende_mesero(mid):
    r    = restaurante_session()
    mesa = Mesa.query.filter_by(id=mid, restaurante_id=r.id).first_or_404()
    mesa.mesero_solicitado = False
    db.session.commit()
    return redirect(url_for("dashboard"))


@app.route("/dashboard/orden/<int:oid>/cancelar", methods=["POST"])
@login_required
def cancelar_orden(oid):
    if session.get("subusuario_id"):
        return redirect(url_for("dashboard"))
    r = restaurante_session()
    o = Orden.query.filter_by(id=oid, restaurante_id=r.id).first_or_404()
    if o.estado in ("pendiente", "confirmada"):
        o.estado = "cancelada"
        db.session.commit()
    return redirect(url_for("dashboard"))


# ══════════════════════════════════════════════
#  MENÚ (gestión de productos)
# ══════════════════════════════════════════════

@app.route("/menu")
@login_required
@plan_requerido
def menu():
    r = restaurante_session()
    productos = Producto.query.filter_by(restaurante_id=r.id).order_by(
        Producto.categoria, Producto.orden_display, Producto.id
    ).all()
    return render_template("restaurante/menu.html", restaurante=r, productos=productos)


@app.route("/menu/agregar", methods=["GET", "POST"])
@login_required
def agregar_producto():
    r = restaurante_session()
    if request.method == "POST":
        nombre      = request.form.get("nombre", "").strip()
        descripcion = request.form.get("descripcion", "").strip()
        categoria   = request.form.get("categoria", "Principal").strip() or "Principal"
        try:
            precio = float(request.form.get("precio", "0"))
        except ValueError:
            flash("Precio inválido.", "error")
            return redirect(url_for("agregar_producto"))

        if not nombre:
            flash("El nombre es requerido.", "error")
            return redirect(url_for("agregar_producto"))

        imagen_url = None
        if "imagen" in request.files and request.files["imagen"].filename:
            try:
                imagen_url = subir_imagen(request.files["imagen"])
            except Exception as e:
                flash(f"Error al subir imagen: {e}", "error")
                return redirect(url_for("agregar_producto"))
        else:
            flash("Debes subir una foto del plato.", "error")
            return redirect(url_for("agregar_producto"))

        p_nuevo = Producto(
            restaurante_id=r.id, nombre=nombre, descripcion=descripcion,
            precio=precio, imagen_url=imagen_url, categoria=categoria,
        )
        db.session.add(p_nuevo)
        db.session.commit()
        flash("Plato creado. Configura sabores aquí si lo necesitas.", "success")
        return redirect(url_for("editar_producto", pid=p_nuevo.id))

    return render_template("restaurante/agregar_producto.html", restaurante=r)


@app.route("/menu/editar/<int:pid>", methods=["GET", "POST"])
@login_required
def editar_producto(pid):
    r = restaurante_session()
    p = Producto.query.filter_by(id=pid, restaurante_id=r.id).first_or_404()

    if request.method == "POST":
        p.nombre      = request.form.get("nombre", p.nombre).strip()
        p.descripcion = request.form.get("descripcion", "").strip()
        p.categoria   = request.form.get("categoria", "Principal").strip() or "Principal"
        p.disponible  = "disponible" in request.form
        try:
            p.precio = float(request.form.get("precio", p.precio))
        except ValueError:
            pass

        if "imagen" in request.files and request.files["imagen"].filename:
            try:
                p.imagen_url = subir_imagen(request.files["imagen"])
            except Exception as e:
                flash(f"Error al subir imagen: {e}", "error")
                return redirect(url_for("editar_producto", pid=pid))

        db.session.commit()
        flash("Plato actualizado.", "success")
        return redirect(url_for("menu"))

    return render_template("restaurante/editar_producto.html", restaurante=r, producto=p)


@app.route("/menu/eliminar/<int:pid>", methods=["POST"])
@login_required
def eliminar_producto(pid):
    r = restaurante_session()
    p = Producto.query.filter_by(id=pid, restaurante_id=r.id).first_or_404()
    if p.items:
        flash("Este plato tiene pedidos anteriores y no puede eliminarse. Usa 'Pausar' para ocultarlo de la carta.", "error")
        return redirect(url_for("menu"))
    db.session.delete(p)
    db.session.commit()
    flash("Plato eliminado.", "success")
    return redirect(url_for("menu"))


@app.route("/menu/toggle/<int:pid>", methods=["POST"])
@login_required
def toggle_producto(pid):
    r = restaurante_session()
    p = Producto.query.filter_by(id=pid, restaurante_id=r.id).first_or_404()
    p.disponible = not p.disponible
    db.session.commit()
    return redirect(url_for("menu"))


@app.route("/menu/producto/<int:pid>/terminos", methods=["POST"])
@login_required
def toggle_terminos_asado(pid):
    r = restaurante_session()
    p = Producto.query.filter_by(id=pid, restaurante_id=r.id).first_or_404()
    p.terminos_asado = not p.terminos_asado
    db.session.commit()
    return redirect(url_for("menu"))


@app.route("/menu/producto/<int:pid>/salsas", methods=["POST"])
@login_required
def toggle_salsas_producto(pid):
    r = restaurante_session()
    p = Producto.query.filter_by(id=pid, restaurante_id=r.id).first_or_404()
    p.salsas_activas = not p.salsas_activas
    db.session.commit()
    return redirect(url_for("menu"))


@app.route("/salsas")
@login_required
def salsas():
    r = restaurante_session()
    lista = Salsa.query.filter_by(restaurante_id=r.id).order_by(Salsa.orden_display).all()
    return render_template("restaurante/salsas.html", restaurante=r, salsas=lista)


@app.route("/salsas/agregar", methods=["POST"])
@login_required
def salsa_agregar():
    r      = restaurante_session()
    nombre = request.form.get("nombre", "").strip()
    if nombre:
        ultimo = Salsa.query.filter_by(restaurante_id=r.id).count()
        db.session.add(Salsa(restaurante_id=r.id, nombre=nombre, orden_display=ultimo))
        db.session.commit()
    return redirect(url_for("salsas"))


@app.route("/salsas/<int:sid>/toggle", methods=["POST"])
@login_required
def salsa_toggle(sid):
    r = restaurante_session()
    s = Salsa.query.filter_by(id=sid, restaurante_id=r.id).first_or_404()
    s.activa = not s.activa
    db.session.commit()
    return redirect(url_for("salsas"))


@app.route("/salsas/<int:sid>/eliminar", methods=["POST"])
@login_required
def salsa_eliminar(sid):
    r = restaurante_session()
    s = Salsa.query.filter_by(id=sid, restaurante_id=r.id).first_or_404()
    db.session.delete(s)
    db.session.commit()
    return redirect(url_for("salsas"))


# ══════════════════════════════════════════════
#  ADICIONES
# ══════════════════════════════════════════════

@app.route("/adiciones")
@login_required
def adiciones():
    r = restaurante_session()
    lista = Adicion.query.filter_by(restaurante_id=r.id).order_by(Adicion.orden_display).all()
    return render_template("restaurante/adiciones.html", restaurante=r, adiciones=lista)


@app.route("/adiciones/agregar", methods=["POST"])
@login_required
def adicion_agregar():
    r = restaurante_session()
    nombre = request.form.get("nombre", "").strip()
    try:
        precio = round(float(request.form.get("precio", 0) or 0), 2)
    except ValueError:
        precio = 0.0
    if nombre:
        orden = Adicion.query.filter_by(restaurante_id=r.id).count()
        db.session.add(Adicion(restaurante_id=r.id, nombre=nombre, precio=precio, orden_display=orden))
        db.session.commit()
    return redirect(url_for("adiciones"))


@app.route("/adiciones/<int:aid>/toggle", methods=["POST"])
@login_required
def adicion_toggle(aid):
    r = restaurante_session()
    a = Adicion.query.filter_by(id=aid, restaurante_id=r.id).first_or_404()
    a.activa = not a.activa
    db.session.commit()
    return redirect(url_for("adiciones"))


@app.route("/adiciones/<int:aid>/eliminar", methods=["POST"])
@login_required
def adicion_eliminar(aid):
    r = restaurante_session()
    a = Adicion.query.filter_by(id=aid, restaurante_id=r.id).first_or_404()
    db.session.delete(a)
    db.session.commit()
    return redirect(url_for("adiciones"))


@app.route("/menu/producto/<int:pid>/adiciones", methods=["POST"])
@login_required
def toggle_adiciones_producto(pid):
    r = restaurante_session()
    p = Producto.query.filter_by(id=pid, restaurante_id=r.id).first_or_404()
    p.adiciones_activas = not p.adiciones_activas
    db.session.commit()
    return redirect(url_for("menu"))


# ══════════════════════════════════════════════
#  BEBIDAS
# ══════════════════════════════════════════════

@app.route("/bebidas")
@login_required
def bebidas():
    r = restaurante_session()
    secciones     = SeccionBebida.query.filter_by(restaurante_id=r.id).order_by(SeccionBebida.orden_display).all()
    return render_template("restaurante/bebidas.html", restaurante=r, secciones=secciones)


@app.route("/bebidas/seccion/agregar", methods=["POST"])
@login_required
def bebida_seccion_agregar():
    r = restaurante_session()
    nombre = request.form.get("nombre", "").strip()
    if nombre:
        orden = SeccionBebida.query.filter_by(restaurante_id=r.id).count()
        db.session.add(SeccionBebida(restaurante_id=r.id, nombre=nombre, orden_display=orden))
        db.session.commit()
    return redirect(url_for("bebidas"))


@app.route("/bebidas/seccion/<int:sid>/toggle", methods=["POST"])
@login_required
def bebida_seccion_toggle(sid):
    r = restaurante_session()
    s = SeccionBebida.query.filter_by(id=sid, restaurante_id=r.id).first_or_404()
    s.activa = not s.activa
    db.session.commit()
    return redirect(url_for("bebidas"))


@app.route("/bebidas/seccion/<int:sid>/eliminar", methods=["POST"])
@login_required
def bebida_seccion_eliminar(sid):
    r = restaurante_session()
    s = SeccionBebida.query.filter_by(id=sid, restaurante_id=r.id).first_or_404()
    db.session.delete(s)
    db.session.commit()
    return redirect(url_for("bebidas"))


@app.route("/bebidas/seccion/<int:sid>/variante/agregar", methods=["POST"])
@login_required
def bebida_variante_agregar(sid):
    r = restaurante_session()
    SeccionBebida.query.filter_by(id=sid, restaurante_id=r.id).first_or_404()
    nombre = request.form.get("nombre", "").strip()
    if nombre:
        orden = VarianteBebida.query.filter_by(seccion_id=sid).count()
        db.session.add(VarianteBebida(seccion_id=sid, nombre=nombre, orden_display=orden))
        db.session.commit()
    return redirect(url_for("bebidas"))


@app.route("/bebidas/variante/<int:vid>/toggle", methods=["POST"])
@login_required
def bebida_variante_toggle(vid):
    r = restaurante_session()
    v = VarianteBebida.query.join(SeccionBebida).filter(
        VarianteBebida.id == vid, SeccionBebida.restaurante_id == r.id).first_or_404()
    v.activa = not v.activa
    db.session.commit()
    return redirect(url_for("bebidas"))


@app.route("/bebidas/variante/<int:vid>/eliminar", methods=["POST"])
@login_required
def bebida_variante_eliminar(vid):
    r = restaurante_session()
    v = VarianteBebida.query.join(SeccionBebida).filter(
        VarianteBebida.id == vid, SeccionBebida.restaurante_id == r.id).first_or_404()
    db.session.delete(v)
    db.session.commit()
    return redirect(url_for("bebidas"))



@app.route("/menu/producto/<int:pid>/sabores/toggle", methods=["POST"])
@login_required
def toggle_sabores_producto(pid):
    r = restaurante_session()
    p = Producto.query.filter_by(id=pid, restaurante_id=r.id).first_or_404()
    p.sabores_activos = not p.sabores_activos
    db.session.commit()
    return redirect(url_for("editar_producto", pid=pid))



@app.route("/menu/producto/<int:pid>/sabor/agregar", methods=["POST"])
@login_required
def agregar_sabor_producto(pid):
    r = restaurante_session()
    p = Producto.query.filter_by(id=pid, restaurante_id=r.id).first_or_404()
    nombre = request.form.get("nombre", "").strip()
    if nombre:
        orden = SaborProducto.query.filter_by(producto_id=pid).count()
        db.session.add(SaborProducto(producto_id=pid, nombre=nombre, orden_display=orden))
        db.session.commit()
    return redirect(url_for("editar_producto", pid=pid))


@app.route("/menu/sabor/<int:sid>/toggle", methods=["POST"])
@login_required
def toggle_sabor_producto(sid):
    r = restaurante_session()
    s = SaborProducto.query.join(Producto).filter(
        SaborProducto.id == sid, Producto.restaurante_id == r.id
    ).first_or_404()
    s.activo = not s.activo
    db.session.commit()
    return redirect(url_for("editar_producto", pid=s.producto_id))


@app.route("/menu/sabor/<int:sid>/eliminar", methods=["POST"])
@login_required
def eliminar_sabor_producto(sid):
    r = restaurante_session()
    s = SaborProducto.query.join(Producto).filter(
        SaborProducto.id == sid, Producto.restaurante_id == r.id
    ).first_or_404()
    pid = s.producto_id
    db.session.delete(s)
    db.session.commit()
    return redirect(url_for("editar_producto", pid=pid))


@app.route("/menu/producto/<int:pid>/bebidas", methods=["POST"])
@login_required
def toggle_bebidas_producto(pid):
    r = restaurante_session()
    p = Producto.query.filter_by(id=pid, restaurante_id=r.id).first_or_404()
    p.bebidas_activas = not p.bebidas_activas
    db.session.commit()
    return redirect(url_for("menu"))


# ══════════════════════════════════════════════
#  MESAS
# ══════════════════════════════════════════════

@app.route("/mesas")
@login_required
def mesas():
    r = restaurante_session()
    if not Mesa.query.filter_by(restaurante_id=r.id, es_para_llevar=True).first():
        db.session.add(Mesa(
            restaurante_id=r.id, numero=0, nombre="Para llevar",
            token=Mesa.nuevo_token(), activa=True, abierta=True, es_para_llevar=True,
        ))
        db.session.commit()
    mesas_list = Mesa.query.filter_by(restaurante_id=r.id).order_by(Mesa.numero).all()
    return render_template("restaurante/mesas.html", restaurante=r, mesas=mesas_list)


@app.route("/mesas/agregar", methods=["POST"])
@login_required
def agregar_mesa():
    r      = restaurante_session()
    nombre = request.form.get("nombre", "").strip()
    ultimo = db.session.query(func.max(Mesa.numero)).filter_by(restaurante_id=r.id).scalar() or 0
    numero = ultimo + 1
    nombre = nombre or f"Mesa {numero}"
    db.session.add(Mesa(restaurante_id=r.id, numero=numero, nombre=nombre, token=Mesa.nuevo_token()))
    db.session.commit()
    flash(f"'{nombre}' creada.", "success")
    return redirect(url_for("mesas"))


@app.route("/mesas/<int:mid>/qr")
@login_required
def imprimir_qr(mid):
    r    = restaurante_session()
    mesa = Mesa.query.filter_by(id=mid, restaurante_id=r.id).first_or_404()
    carta_url = request.host_url.rstrip("/") + f"/carta/{r.slug}/{mesa.token}"
    return render_template("restaurante/qr_imprimir.html",
        restaurante=r, mesa=mesa, carta_url=carta_url)


@app.route("/mesas/<int:mid>/toggle-sesion", methods=["POST"])
@login_required
def toggle_sesion_mesa(mid):
    r = restaurante_session()
    m = Mesa.query.filter_by(id=mid, restaurante_id=r.id).first_or_404()
    m.abierta = not m.abierta
    db.session.commit()
    estado = "abierta" if m.abierta else "cerrada"
    flash(f"{m.nombre} {estado}.", "success")
    return redirect(url_for("mesas"))


@app.route("/mesas/reiniciar", methods=["POST"])
@login_required
def reiniciar_mesas():
    r = restaurante_session()
    mesas_list = Mesa.query.filter_by(restaurante_id=r.id).all()
    for m in mesas_list:
        ordenes = Orden.query.filter_by(mesa_id=m.id).all()
        for o in ordenes:
            db.session.delete(o)
        if not m.es_para_llevar:
            db.session.delete(m)
    db.session.commit()
    flash("Todo reiniciado. Órdenes y mesas eliminadas.", "success")
    return redirect(url_for("mesas"))


@app.route("/mesas/eliminar/<int:mid>", methods=["POST"])
@login_required
def eliminar_mesa(mid):
    r = restaurante_session()
    m = Mesa.query.filter_by(id=mid, restaurante_id=r.id).first_or_404()
    if m.es_para_llevar:
        flash("La mesa 'Para llevar' no se puede eliminar.", "error")
        return redirect(url_for("mesas"))
    if Orden.query.filter_by(mesa_id=m.id).first():
        flash(f"No se puede eliminar '{m.nombre}' porque tiene historial de pedidos.", "error")
        return redirect(url_for("mesas"))
    try:
        db.session.delete(m)
        db.session.commit()
        flash("Mesa eliminada.", "success")
    except Exception:
        db.session.rollback()
        flash(f"No se puede eliminar '{m.nombre}' porque tiene pedidos asociados.", "error")
    return redirect(url_for("mesas"))


# ══════════════════════════════════════════════
#  CARTA PÚBLICA (lo que ve el cliente)
# ══════════════════════════════════════════════

@app.route("/carta/<slug>/<mesa_token>")
def carta(slug, mesa_token):
    r    = Restaurante.query.filter_by(slug=slug, activo=True).first_or_404()
    mesa = Mesa.query.filter_by(token=mesa_token, restaurante_id=r.id, activa=True).first_or_404()

    # Verificar que el cliente esté en la misma red WiFi que el restaurante
    if r.restringir_red and r.ip_red and r.email != "demo@cartadigital.app":
        cliente_ip = get_client_ip()
        if cliente_ip != r.ip_red:
            return render_template("carta/red_requerida.html", restaurante=r)

    # En bares la mesa siempre está abierta para nuevos pedidos hasta que el admin cierre la cuenta
    if r.categoria == 'bar':
        if not mesa.abierta:
            mesa.abierta   = True
            mesa.tab_inicio = datetime.utcnow()
            db.session.commit()
    else:
        orden_activa = Orden.query.filter(
            Orden.mesa_id == mesa.id,
            Orden.estado.in_(["pendiente", "confirmada", "lista"]),
        ).first()
        if not mesa.abierta and not orden_activa:
            mesa.abierta = True
            db.session.commit()

    productos = Producto.query.filter_by(
        restaurante_id=r.id, disponible=True
    ).order_by(Producto.categoria, Producto.orden_display).all()

    categorias, cat_map = [], {}
    for p in productos:
        if p.categoria not in cat_map:
            categorias.append(p.categoria)
            cat_map[p.categoria] = []
        cat_map[p.categoria].append(p)

    carrito = session.get(f"cart_{mesa_token}", {})
    total_carrito  = sum(v["precio"] * v["cantidad"] for v in carrito.values())
    items_carrito  = sum(v["cantidad"] for v in carrito.values())

    return render_template("carta/index.html",
        restaurante=r, mesa=mesa,
        categorias=categorias, cat_map=cat_map,
        carrito=carrito, total_carrito=total_carrito,
        items_carrito=items_carrito, max_cantidad=MAX_CANTIDAD,
        mesa_abierta=mesa.abierta,
        terminos_asado=TERMINOS_ASADO,
        salsas_rest=Salsa.query.filter_by(restaurante_id=r.id, activa=True).order_by(Salsa.orden_display).all(),
        adiciones_rest=Adicion.query.filter_by(restaurante_id=r.id, activa=True).order_by(Adicion.orden_display).all(),
        secciones_bebida=SeccionBebida.query.filter_by(restaurante_id=r.id, activa=True).order_by(SeccionBebida.orden_display).all(),
        modo_cobro=r.modo_cobro,
        metodos_pago_carta=MetodoPago.query.filter_by(restaurante_id=r.id, activo=True).order_by(MetodoPago.orden_display).all() if r.modo_cobro else [],
    )


@app.route("/carta/<slug>/<mesa_token>/agregar", methods=["POST"])
def carta_agregar(slug, mesa_token):
    r    = Restaurante.query.filter_by(slug=slug, activo=True).first_or_404()
    mesa = Mesa.query.filter_by(token=mesa_token, restaurante_id=r.id, activa=True).first_or_404()
    if not mesa.abierta:
        return redirect(url_for("carta", slug=slug, mesa_token=mesa_token))

    prod_id  = request.form.get("producto_id", type=int)
    cantidad = max(1, min(request.form.get("cantidad", 1, type=int), MAX_CANTIDAD))
    p        = Producto.query.filter_by(id=prod_id, restaurante_id=r.id, disponible=True).first_or_404()

    termino_asado  = request.form.get("termino_asado",  "").strip()
    termino_salsa  = request.form.get("termino_salsa",  "").strip()
    termino_bebida = request.form.get("termino_bebida", "").strip()

    # Sabor por producto (per-product flat list)
    termino_sabor = ""
    if p.sabores_activos:
        sabor_raw = request.form.get("termino_sabor", "").strip()
        if sabor_raw:
            sabor_valido = SaborProducto.query.filter_by(
                producto_id=p.id, nombre=sabor_raw, activo=True
            ).first()
            if sabor_valido:
                termino_sabor = sabor_raw

    if p.terminos_asado and termino_asado not in TERMINOS_ASADO:
        return redirect(url_for("carta", slug=slug, mesa_token=mesa_token))

    # Adiciones (checkboxes — múltiples)
    adicion_ids_raw = request.form.getlist("adicion_id")
    extra_precio    = 0.0
    adicion_partes  = []
    adicion_key     = ""
    if p.adiciones_activas and adicion_ids_raw:
        ids_validos = [int(i) for i in adicion_ids_raw if i.isdigit()]
        if ids_validos:
            adic_objs = Adicion.query.filter(
                Adicion.id.in_(ids_validos),
                Adicion.restaurante_id == r.id,
                Adicion.activa == True,
            ).all()
            for a in adic_objs:
                adicion_partes.append(f"+{a.nombre}")
                extra_precio += a.precio
            adicion_key = ",".join(sorted(str(a.id) for a in adic_objs))

    partes  = [t for t in [termino_asado, termino_salsa, termino_bebida, termino_sabor] if t] + adicion_partes
    termino = " · ".join(partes)

    precio_final = round(p.precio + extra_precio, 2)

    cart_key  = f"cart_{mesa_token}"
    carrito   = session.get(cart_key, {})
    key_parts = [str(prod_id)]
    if termino:    key_parts.append(termino)
    if adicion_key: key_parts.append(adicion_key)
    item_key  = "_".join(key_parts)
    actual    = carrito.get(item_key, {}).get("cantidad", 0)

    etiqueta = f"{p.nombre} ({termino})" if termino else p.nombre
    carrito[item_key] = {
        "nombre":   etiqueta,
        "precio":   precio_final,
        "cantidad": min(actual + cantidad, MAX_CANTIDAD),
        "imagen":   p.imagen_url or "",
        "termino":  termino,
    }
    session[cart_key] = carrito
    return redirect(url_for("carta", slug=slug, mesa_token=mesa_token) + "#carrito")


@app.route("/carta/<slug>/<mesa_token>/llamar-mesero", methods=["POST"])
def llamar_mesero(slug, mesa_token):
    mesa = Mesa.query.filter_by(token=mesa_token).first_or_404()
    if not mesa.es_para_llevar:
        mesa.mesero_solicitado = True
        db.session.commit()
    return {"ok": True}, 200


@app.route("/carta/<slug>/<mesa_token>/cambiar", methods=["POST"])
def carta_cambiar(slug, mesa_token):
    cart_key = f"cart_{mesa_token}"
    carrito  = session.get(cart_key, {})
    item_key = request.form.get("item_key", "")
    accion   = request.form.get("accion")  # "mas" | "menos" | "quitar"

    if item_key in carrito:
        if accion == "mas":
            carrito[item_key]["cantidad"] = min(carrito[item_key]["cantidad"] + 1, MAX_CANTIDAD)
        elif accion == "menos":
            carrito[item_key]["cantidad"] -= 1
            if carrito[item_key]["cantidad"] <= 0:
                del carrito[item_key]
        elif accion == "quitar":
            del carrito[item_key]

    session[cart_key] = carrito
    return redirect(url_for("carta", slug=slug, mesa_token=mesa_token) + "#carrito")


@app.route("/carta/<slug>/<mesa_token>/pedido", methods=["POST"])
def hacer_pedido(slug, mesa_token):
    r    = Restaurante.query.filter_by(slug=slug, activo=True).first_or_404()
    mesa = Mesa.query.filter_by(token=mesa_token, restaurante_id=r.id, activa=True).first_or_404()

    if not mesa.abierta:
        flash("Esta mesa no está disponible.", "error")
        return redirect(url_for("carta", slug=slug, mesa_token=mesa_token))

    cart_key = f"cart_{mesa_token}"
    carrito  = session.get(cart_key, {})
    if not carrito:
        flash("Tu carrito está vacío.", "error")
        return redirect(url_for("carta", slug=slug, mesa_token=mesa_token))

    nombre_cliente = request.form.get("nombre_cliente", "").strip()
    if mesa.es_para_llevar and not nombre_cliente:
        flash("Por favor ingresa tu nombre para el pedido para llevar.", "error")
        return redirect(url_for("carta", slug=slug, mesa_token=mesa_token))
    nombre_cliente = nombre_cliente or "Cliente"
    notas          = request.form.get("notas", "").strip()

    total, items = 0.0, []
    for item_key, item in carrito.items():
        prod_id = int(item_key.split("_")[0])
        p = Producto.query.filter_by(id=prod_id, restaurante_id=r.id, disponible=True).first()
        if not p:
            continue
        cant        = max(1, min(item["cantidad"], MAX_CANTIDAD))
        precio_unit = round(item.get("precio", p.precio), 2)
        subtotal    = round(precio_unit * cant, 2)
        total      += subtotal
        termino     = item.get("termino", "")
        items.append(ItemOrden(
            producto_id=p.id, cantidad=cant,
            precio_unitario=precio_unit, subtotal=subtotal,
            notas_item=termino if termino else None,
        ))

    if not items:
        flash("No hay productos disponibles en tu pedido.", "error")
        return redirect(url_for("carta", slug=slug, mesa_token=mesa_token))

    orden = Orden(
        restaurante_id=r.id, mesa_id=mesa.id,
        token=Orden.nuevo_token(),
        nombre_cliente=nombre_cliente,
        total=round(total, 2), notas=notas,
    )
    if r.modo_cobro:
        metodo = request.form.get("metodo_preferido", "").strip()
        orden.solicita_cuenta  = True
        orden.metodo_preferido = metodo if metodo else None

    db.session.add(orden)
    db.session.flush()
    for item in items:
        item.orden_id = orden.id
        db.session.add(item)
    db.session.commit()

    session.pop(cart_key, None)
    return redirect(url_for("estado_orden", token=orden.token))


# ══════════════════════════════════════════════
#  SEGUIMIENTO DE ORDEN (cliente)
# ══════════════════════════════════════════════

@app.route("/orden/<token>")
def estado_orden(token):
    orden = Orden.query.filter_by(token=token).first_or_404()
    if not orden.restaurante.activo or not plan_vigente(orden.restaurante):
        return render_template("auth/plan_vencido_cliente.html", restaurante=orden.restaurante), 403
    metodos_pago = MetodoPago.query.filter_by(
        restaurante_id=orden.restaurante_id, activo=True
    ).order_by(MetodoPago.orden_display).all()
    pedidos_adelante = Orden.query.filter(
        Orden.restaurante_id == orden.restaurante_id,
        Orden.estado == "pendiente",
        Orden.id < orden.id,
    ).count()
    return render_template("carta/orden.html", orden=orden,
                           restaurante=orden.restaurante, metodos_pago=metodos_pago,
                           modo_cobro=orden.restaurante.modo_cobro,
                           pedidos_adelante=pedidos_adelante)


@app.route("/api/orden/<token>/estado")
def api_estado_orden(token):
    orden = Orden.query.filter_by(token=token).first_or_404()
    if not orden.restaurante.activo or not plan_vigente(orden.restaurante):
        return jsonify({"error": "no_disponible"}), 403
    pedidos_adelante = Orden.query.filter(
        Orden.restaurante_id == orden.restaurante_id,
        Orden.estado == "pendiente",
        Orden.id < orden.id,
    ).count() if orden.estado == "pendiente" else 0
    return jsonify({
        "estado":           orden.estado,
        "mesa_abierta":     orden.mesa.abierta if orden.mesa else False,
        "pedidos_adelante": pedidos_adelante,
    })


@app.route("/orden/<token>/cuenta", methods=["POST"])
def solicitar_cuenta(token):
    orden = Orden.query.filter_by(token=token).first_or_404()
    if not orden.restaurante.activo or not plan_vigente(orden.restaurante):
        return render_template("auth/plan_vencido_cliente.html", restaurante=orden.restaurante), 403
    if orden.estado == "lista":
        orden.solicita_cuenta  = True
        orden.metodo_preferido = request.form.get("metodo_preferido", "efectivo")
        db.session.commit()
    return redirect(url_for("estado_orden", token=token))


@app.route("/orden/<token>/recibo")
def recibo_orden(token):
    orden = Orden.query.filter_by(token=token).first_or_404()
    return render_template("carta/recibo.html", orden=orden, restaurante=orden.restaurante)


# ══════════════════════════════════════════════
#  REPORTES
# ══════════════════════════════════════════════

@app.route("/guia")
@login_required
def guia():
    return render_template("restaurante/guia.html")


@app.route("/reportes")
@login_required
@plan_requerido
def reportes():
    r   = restaurante_session()
    hoy = date.today()

    # Resumen mensual de los últimos 12 meses
    resumen_mensual = db.session.query(
        func.extract("year",  Orden.fecha).label("anio"),
        func.extract("month", Orden.fecha).label("mes"),
        func.count(Orden.id).label("ordenes"),
        func.sum(Orden.total).label("total"),
    ).filter(
        Orden.restaurante_id == r.id,
        Orden.estado == "pagada",
    ).group_by("anio", "mes").order_by("anio", "mes").all()

    return render_template("restaurante/reportes.html",
        restaurante=r, hoy=hoy, resumen_mensual=resumen_mensual)


@app.route("/reportes/descargar")
@login_required
def descargar_reporte():
    r         = restaurante_session()
    tipo      = request.args.get("tipo", "diario")
    fecha_str = request.args.get("fecha", date.today().isoformat())

    try:
        fecha_ref = date.fromisoformat(fecha_str)
    except ValueError:
        fecha_ref = date.today()

    if tipo == "diario":
        # Convertir el día Ecuador a rango UTC (UTC-5 → sumar 5h)
        inicio_utc = datetime.combine(fecha_ref, time(0, 0, 0)) - EC_OFFSET
        fin_utc    = datetime.combine(fecha_ref, time(23, 59, 59)) - EC_OFFSET
        ordenes = Orden.query.filter(
            Orden.restaurante_id == r.id,
            Orden.estado == "pagada",
            Orden.fecha >= inicio_utc,
            Orden.fecha <= fin_utc,
        ).order_by(Orden.fecha).all()
        nombre_archivo = f"reporte_diario_{fecha_ref}.xlsx"
    else:
        # Primer y último instante del mes en UTC
        primer_dia  = fecha_ref.replace(day=1)
        if fecha_ref.month == 12:
            ultimo_dia = fecha_ref.replace(year=fecha_ref.year + 1, month=1, day=1)
        else:
            ultimo_dia = fecha_ref.replace(month=fecha_ref.month + 1, day=1)
        inicio_utc = datetime.combine(primer_dia, time(0, 0, 0)) - EC_OFFSET
        fin_utc    = datetime.combine(ultimo_dia, time(0, 0, 0)) - EC_OFFSET
        ordenes = Orden.query.filter(
            Orden.restaurante_id == r.id,
            Orden.estado == "pagada",
            Orden.fecha >= inicio_utc,
            Orden.fecha < fin_utc,
        ).order_by(Orden.fecha).all()
        nombre_archivo = f"reporte_mensual_{fecha_ref.year}_{fecha_ref.month:02d}.xlsx"

    # ── Estilos ──────────────────────────────────────────────────────────
    naranja     = "F97316"
    gris_fondo  = "F8FAFC"
    borde_color = "E2E8F0"

    def borde_fino():
        s = Side(style="thin", color=borde_color)
        return Border(left=s, right=s, top=s, bottom=s)

    # ── Libro Excel ──────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    # ── Productos únicos del período (orden alfabético) ───────────────────
    productos_map = {}
    for o in ordenes:
        for item in o.items:
            if item.producto and item.producto_id not in productos_map:
                productos_map[item.producto_id] = item.producto.nombre
    prod_ids = sorted(productos_map, key=lambda pid: productos_map[pid])

    # ── Cabecera ─────────────────────────────────────────────────────────
    cols_fijas  = ["Fecha", "Hora", "# Orden", "Mesa", "Cliente"]
    cols_prod   = [productos_map[pid] for pid in prod_ids]
    cols_fin    = ["TOTAL ($)", "Método de pago"]
    cabecera    = cols_fijas + cols_prod + cols_fin

    for col_idx, titulo in enumerate(cabecera, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor=naranja)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = borde_fino()
    ws.row_dimensions[1].height = 30

    # ── Filas de datos ────────────────────────────────────────────────────
    totales_qty  = {pid: 0 for pid in prod_ids}
    total_global = 0.0
    fila_excel   = 2

    for o in ordenes:
        ec_fecha = o.fecha + EC_OFFSET

        item_map = {}
        for item in o.items:
            if item.producto:
                if item.producto_id not in item_map:
                    item_map[item.producto_id] = {"qty": 0, "notas": []}
                item_map[item.producto_id]["qty"] += item.cantidad
                if item.notas_item:
                    item_map[item.producto_id]["notas"].append(item.notas_item)

        valores = [
            ec_fecha.strftime("%d/%m/%Y"),
            ec_fecha.strftime("%H:%M"),
            o.id,
            o.mesa.nombre if o.mesa else "—",
            o.nombre_cliente or "—",
        ]
        for pid in prod_ids:
            if pid in item_map:
                d = item_map[pid]
                val = d["qty"]
                if d["notas"]:
                    val = f"{d['qty']} ({', '.join(d['notas'])})"
                valores.append(val)
                totales_qty[pid] += d["qty"]
            else:
                valores.append("")
        valores.append(round(o.total, 2))
        valores.append(o.metodo_pago or "—")
        total_global += o.total

        bg = "FFFFFF" if fila_excel % 2 == 0 else gris_fondo
        for col_idx, val in enumerate(valores, start=1):
            cell = ws.cell(row=fila_excel, column=col_idx, value=val)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.border    = borde_fino()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font      = Font(size=10)
        fila_excel += 1

    # ── Fila de totales ───────────────────────────────────────────────────
    fila_excel += 1  # fila en blanco
    fila_tot = fila_excel
    totales_vals = ["TOTAL", "", "", "", ""] \
                 + [totales_qty[pid] if totales_qty[pid] > 0 else "" for pid in prod_ids] \
                 + [round(total_global, 2), ""]

    for col_idx, val in enumerate(totales_vals, start=1):
        cell = ws.cell(row=fila_tot, column=col_idx, value=val)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = PatternFill("solid", fgColor=naranja)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = borde_fino()
    ws.row_dimensions[fila_tot].height = 22

    # ── Ancho de columnas automático ─────────────────────────────────────
    anchos_min = {"A": 12, "B": 8, "C": 9, "D": 12, "E": 14}
    for col_idx, titulo in enumerate(cabecera, start=1):
        letra = get_column_letter(col_idx)
        ancho = max(len(str(titulo)) + 2, anchos_min.get(letra, 10))
        ws.column_dimensions[letra].width = min(ancho, 30)

    # ── Respuesta ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre_archivo = nombre_archivo.replace(".csv", ".xlsx")
    resp = make_response(buf.read())
    resp.headers["Content-Disposition"] = f"attachment; filename={nombre_archivo}"
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return resp


# ══════════════════════════════════════════════
#  PERFIL
# ══════════════════════════════════════════════

@app.route("/perfil", methods=["GET", "POST"])
@login_required
@plan_requerido
def perfil():
    r = restaurante_session()
    if request.method == "POST":
        r.nombre        = request.form.get("nombre", r.nombre).strip()
        r.whatsapp      = request.form.get("whatsapp", "").strip()
        r.ciudad        = request.form.get("ciudad", "").strip()
        r.descripcion   = request.form.get("descripcion", "").strip()
        r.restringir_red = "restringir_red" in request.form

        pwd = request.form.get("password", "")
        if pwd:
            pwd_actual = request.form.get("password_actual", "")
            if not r.check_password(pwd_actual):
                flash("La contraseña actual es incorrecta.", "error")
                return redirect(url_for("perfil"))
            if len(pwd) < 6:
                flash("La contraseña debe tener al menos 6 caracteres.", "error")
                return redirect(url_for("perfil"))
            r.set_password(pwd)

        if "logo" in request.files and request.files["logo"].filename:
            try:
                r.logo_url = subir_imagen(request.files["logo"], "carta_logos")
            except Exception as e:
                flash(f"Error al subir logo: {e}", "error")
                return redirect(url_for("perfil"))

        db.session.commit()
        flash("Perfil actualizado.", "success")
        return redirect(url_for("perfil"))

    return render_template("restaurante/perfil.html", restaurante=r, ip_actual=get_client_ip())


@app.route("/perfil/actualizar-ip", methods=["POST"])
@login_required
def actualizar_ip_red():
    r = restaurante_session()
    r.ip_red = get_client_ip()
    db.session.commit()
    flash(f"IP de red actualizada a {r.ip_red}. Los clientes en este WiFi podrán ver el menú.", "success")
    return redirect(url_for("perfil"))


@app.route("/perfil/modo-cobro", methods=["POST"])
@login_required
def toggle_modo_cobro():
    r = restaurante_session()
    r.modo_cobro = not r.modo_cobro
    db.session.commit()
    return redirect(url_for("perfil"))


DEMO_EMAIL = "demo@cartadigital.app"
DEMO_PASS  = "demo1234"

DEMO_PRODUCTOS = [
    ("Ceviche de camarón",   "Camarones frescos marinados en limón, cebolla morada, cilantro y ají. Servido con chifles y patacones.",  7.50, "Entradas",  "https://picsum.photos/seed/ceviche1/900/700"),
    ("Patacones con hogao",  "Patacones crocantes acompañados de hogao casero y queso costeño.",                                        4.50, "Entradas",  "https://picsum.photos/seed/patacon2/900/700"),
    ("Seco de pollo",        "Pollo en salsa criolla de cerveza y especias, servido con arroz, menestra y aguacate.",                    8.50, "Principal", "https://picsum.photos/seed/chicken3/900/700"),
    ("Bandeja paisa",        "Frijoles, chicharrón, chorizo, huevo frito, arroz, aguacate, maduro y arepa. El plato más completo.",     10.00, "Principal", "https://picsum.photos/seed/bandeja4/900/700"),
    ("Arroz con mariscos",   "Arroz marinero con camarones, calamar y mejillones en salsa de tomate y azafrán.",                        12.00, "Principal", "https://picsum.photos/seed/seafood5/900/700"),
    ("Churrasco a la parrilla","350g de res a la parrilla con chimichurri casero, papas fritas y ensalada fresca.",                    14.00, "Principal", "https://picsum.photos/seed/steak6/900/700"),
    ("Sopa de quinua",       "Sopa andina con quinua, papa, zanahoria, apio y hierbas aromáticas. Reconfortante y nutritiva.",          5.50, "Sopas",     "https://picsum.photos/seed/soup7/900/700"),
    ("Caldo de gallina",     "Caldo tradicional de gallina criolla con papa, yuca, hierbas y ají.",                                     5.00, "Sopas",     "https://picsum.photos/seed/broth8/900/700"),
    ("Tres leches",          "Bizcocho bañado en tres tipos de leche, cubierto con crema chantilly y canela.",                          4.00, "Postres",   "https://picsum.photos/seed/cake9/900/700"),
    ("Flan de coco",         "Flan artesanal de coco con caramelo natural. Cremoso y ligero.",                                          3.50, "Postres",   "https://picsum.photos/seed/flan10/900/700"),
    ("Jugo de naranja",      "Naranja recién exprimida, 100% natural. Sin azúcar añadida.",                                             2.50, "Bebidas",   "https://picsum.photos/seed/juice11/900/700"),
    ("Cola de avena",        "Bebida tradicional de avena con leche, canela y azúcar. Fría y refrescante.",                             2.00, "Bebidas",   "https://picsum.photos/seed/oat12/900/700"),
    ("Cerveza artesanal",    "Cerveza artesanal local tipo lager. Fría y refrescante.",                                                  3.50, "Bebidas",   "https://picsum.photos/seed/beer13/900/700"),
]


def crear_demo():
    r = Restaurante.query.filter_by(email=DEMO_EMAIL).first()
    if not r:
        slug = "restaurante-demo"
        r = Restaurante(
            nombre="El Rincón Criollo",
            email=DEMO_EMAIL,
            slug=slug,
            whatsapp="0991234567",
            ciudad="Quito",
            descripcion="Cocina criolla ecuatoriana · Lun–Dom 8:00–22:00",
            logo_url="https://picsum.photos/seed/logorest/200/200",
        )
        r.set_password(DEMO_PASS)
        r.plan       = 'anual'
        r.plan_vence = datetime(2099, 12, 31)
        db.session.add(r)
        db.session.flush()

        for i in range(1, 9):
            db.session.add(Mesa(restaurante_id=r.id, numero=i,
                                nombre=f"Mesa {i}", token=Mesa.nuevo_token()))

        for i, (nombre, desc, precio, cat, img) in enumerate(DEMO_PRODUCTOS):
            db.session.add(Producto(
                restaurante_id=r.id, nombre=nombre, descripcion=desc,
                precio=precio, categoria=cat, imagen_url=img,
                disponible=True, orden_display=i,
            ))
        db.session.commit()
    return r


@app.route("/demo")
def demo_landing():
    r    = crear_demo()
    mesa = Mesa.query.filter_by(restaurante_id=r.id, numero=1).first()
    carta_url = url_for("carta", slug=r.slug, mesa_token=mesa.token, _external=False)
    return render_template("demo.html", restaurante=r, carta_url=carta_url)


@app.route("/demo/admin")
def demo_admin():
    r = crear_demo()
    session["restaurante_id"] = r.id
    flash("Estás en el panel de demostración — explora libremente.", "info")
    return redirect(url_for("dashboard"))


@app.route("/demo/prueba")
def demo_prueba():
    r    = crear_demo()
    session["restaurante_id"] = r.id          # auto-login como demo admin
    mesa = Mesa.query.filter_by(restaurante_id=r.id, numero=1).first()
    carta_url    = url_for("carta",     slug=r.slug, mesa_token=mesa.token)
    dashboard_url = url_for("dashboard")
    return render_template("demo_prueba.html",
        restaurante=r, carta_url=carta_url, dashboard_url=dashboard_url)


@app.route("/api/meseros-solicitados")
@login_required
def api_meseros_solicitados():
    from flask import jsonify
    r = restaurante_session()
    mesas = Mesa.query.filter_by(restaurante_id=r.id, mesero_solicitado=True).all()
    return jsonify({
        "count": len(mesas),
        "mesas": [{"id": m.id, "nombre": m.nombre} for m in mesas],
    })


@app.route("/api/ordenes-activas")
@login_required
def api_ordenes_activas():
    from flask import jsonify
    r = restaurante_session()
    count = Orden.query.filter(
        Orden.restaurante_id == r.id,
        Orden.estado.in_(["pendiente", "confirmada", "lista"])
    ).count()
    pendientes = Orden.query.filter(
        Orden.restaurante_id == r.id,
        Orden.estado == "pendiente"
    ).count()
    cuentas = Orden.query.filter(
        Orden.restaurante_id == r.id,
        Orden.estado == "lista",
        Orden.solicita_cuenta == True,
    ).count()
    return jsonify({"count": count, "pendientes": pendientes, "cuentas": cuentas})


# ══════════════════════════════════════════════
#  MÉTODOS DE PAGO
# ══════════════════════════════════════════════

@app.route("/metodos-pago")
@login_required
def metodos_pago():
    r = restaurante_session()
    metodos = MetodoPago.query.filter_by(restaurante_id=r.id).order_by(MetodoPago.orden_display).all()
    return render_template("restaurante/metodos_pago.html", restaurante=r, metodos=metodos)


@app.route("/metodos-pago/agregar", methods=["POST"])
@login_required
def agregar_metodo_pago():
    r      = restaurante_session()
    nombre = request.form.get("nombre", "").strip()
    icono  = request.form.get("icono",  "💳").strip() or "💳"
    if nombre:
        ultimo = db.session.query(func.max(MetodoPago.orden_display)).filter_by(restaurante_id=r.id).scalar() or 0
        db.session.add(MetodoPago(restaurante_id=r.id, nombre=nombre, icono=icono, orden_display=ultimo+1))
        db.session.commit()
        flash(f"Método '{nombre}' agregado.", "success")
    return redirect(url_for("metodos_pago"))


@app.route("/metodos-pago/<int:mid>/toggle", methods=["POST"])
@login_required
def toggle_metodo_pago(mid):
    r = restaurante_session()
    m = MetodoPago.query.filter_by(id=mid, restaurante_id=r.id).first_or_404()
    m.activo = not m.activo
    db.session.commit()
    return redirect(url_for("metodos_pago"))


@app.route("/metodos-pago/<int:mid>/eliminar", methods=["POST"])
@login_required
def eliminar_metodo_pago(mid):
    r = restaurante_session()
    m = MetodoPago.query.filter_by(id=mid, restaurante_id=r.id).first_or_404()
    db.session.delete(m)
    db.session.commit()
    flash("Método eliminado.", "success")
    return redirect(url_for("metodos_pago"))


# ══════════════════════════════════════════════
#  SOPORTE
# ══════════════════════════════════════════════

@app.route("/soporte/enviar", methods=["POST"])
def soporte_enviar():
    from flask import jsonify
    nombre  = request.form.get("nombre",  "").strip()
    email   = request.form.get("email",   "").strip()
    mensaje = request.form.get("mensaje", "").strip()
    if not mensaje:
        return jsonify({"ok": False, "error": "El mensaje no puede estar vacío"}), 400
    db.session.add(MensajeSoporte(nombre=nombre, email=email, mensaje=mensaje))
    db.session.commit()
    return jsonify({"ok": True})


# ══════════════════════════════════════════════
#  SUPER-ADMIN
# ══════════════════════════════════════════════

ADMIN_USER = os.getenv("ADMIN_USER", "admin")
ADMIN_PASS = os.getenv("ADMIN_PASSWORD", "")


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("is_admin"):
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return decorated


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        usuario  = request.form.get("usuario", "")
        password = request.form.get("password", "")
        if usuario == ADMIN_USER:
            cfg = ConfigGlobal.query.filter_by(clave="admin_password_hash").first()
            ok  = check_password_hash(cfg.valor, password) if cfg else (password == ADMIN_PASS)
            if ok:
                session["is_admin"] = True
                return redirect(url_for("admin_panel"))
        return render_template("admin/login.html", error="Credenciales incorrectas")
    return render_template("admin/login.html", error=None)


@app.route("/admin/cambiar-password", methods=["POST"])
@admin_required
def admin_cambiar_password():
    pwd_actual = request.form.get("password_actual", "")
    pwd        = request.form.get("password", "")
    pwd2       = request.form.get("password2", "")
    # Verificar contraseña actual
    cfg = ConfigGlobal.query.filter_by(clave="admin_password_hash").first()
    ok  = check_password_hash(cfg.valor, pwd_actual) if cfg else (pwd_actual == ADMIN_PASS)
    if not ok:
        flash("La contraseña actual es incorrecta.", "error")
    elif pwd != pwd2:
        flash("Las contraseñas no coinciden.", "error")
    elif len(pwd) < 6:
        flash("La contraseña debe tener al menos 6 caracteres.", "error")
    else:
        if not cfg:
            cfg = ConfigGlobal(clave="admin_password_hash", valor="")
            db.session.add(cfg)
        cfg.valor = generate_password_hash(pwd)
        db.session.commit()
        flash("Contraseña de administrador actualizada.", "success")
    return redirect(url_for("admin_panel"))


@app.route("/admin/logout")
def admin_logout():
    session.pop("is_admin", None)
    return redirect(url_for("admin_login"))


@app.route("/admin")
@admin_required
def admin_panel():
    restaurantes  = Restaurante.query.order_by(Restaurante.fecha_registro.desc()).all()
    msgs_no_leidos = MensajeSoporte.query.filter_by(leido=False).count()

    stats = {}
    for r in restaurantes:
        total_ordenes  = Orden.query.filter_by(restaurante_id=r.id).count()
        ordenes_hoy    = Orden.query.filter(
            Orden.restaurante_id == r.id,
            func.date(Orden.fecha) == date.today(),
        ).count()
        ingresos_total = db.session.query(func.sum(Orden.total)).filter(
            Orden.restaurante_id == r.id,
            Orden.estado == "pagada",
        ).scalar() or 0.0
        stats[r.id] = {
            "total_ordenes":  total_ordenes,
            "ordenes_hoy":    ordenes_hoy,
            "ingresos_total": ingresos_total,
        }

    return render_template("admin/panel.html", restaurantes=restaurantes, stats=stats,
                           msgs_no_leidos=msgs_no_leidos)


@app.route("/admin/restaurante/<int:rid>/panel")
@admin_required
def admin_ver_panel(rid):
    r = Restaurante.query.get_or_404(rid)
    session["restaurante_id"] = r.id   # impersona al restaurante
    return redirect(url_for("dashboard"))


@app.route("/admin/restaurante/<int:rid>/verificar", methods=["POST"])
@admin_required
def admin_verificar_email(rid):
    r = Restaurante.query.get_or_404(rid)
    r.email_verificado = True
    db.session.commit()
    flash(f"Email de {r.nombre} verificado manualmente.", "success")
    return redirect(url_for("admin_panel"))


@app.route("/admin/restaurante/<int:rid>/toggle", methods=["POST"])
@admin_required
def admin_toggle_restaurante(rid):
    r = Restaurante.query.get_or_404(rid)
    r.activo = not r.activo
    db.session.commit()
    return redirect(url_for("admin_panel"))


@app.route("/admin/restaurante/<int:rid>/eliminar", methods=["POST"])
@admin_required
def admin_eliminar_restaurante(rid):
    r = Restaurante.query.get_or_404(rid)
    db.session.delete(r)
    db.session.commit()
    return redirect(url_for("admin_panel"))


@app.route("/admin/mensajes")
@admin_required
def admin_mensajes():
    mensajes = MensajeSoporte.query.order_by(MensajeSoporte.fecha.desc()).all()
    # marcar todos como leídos al abrir la página
    MensajeSoporte.query.filter_by(leido=False).update({"leido": True})
    db.session.commit()
    return render_template("admin/mensajes.html", mensajes=mensajes)


@app.route("/admin/mensajes/<int:mid>/eliminar", methods=["POST"])
@admin_required
def admin_eliminar_mensaje(mid):
    m = MensajeSoporte.query.get_or_404(mid)
    db.session.delete(m)
    db.session.commit()
    return redirect(url_for("admin_mensajes"))


@app.route("/admin/restaurante/<int:rid>/plan", methods=["POST"])
@admin_required
def admin_set_plan(rid):
    r         = Restaurante.query.get_or_404(rid)
    plan_tipo = request.form.get("plan", "")
    fecha_str = request.form.get("fecha_inicio", "").strip()

    if plan_tipo not in PLANES and plan_tipo != "trial":
        flash("Plan inválido.", "error")
        return redirect(url_for("admin_panel"))

    try:
        fecha_inicio = datetime.strptime(fecha_str, "%Y-%m-%d") if fecha_str else datetime.utcnow()
    except ValueError:
        fecha_inicio = datetime.utcnow()

    r.plan        = plan_tipo
    r.plan_inicio = fecha_inicio
    if plan_tipo == "trial":
        dias = TRIAL_DIAS_POR_PAIS.get(r.pais or 'ecuador', TRIAL_DIAS)
        r.plan_vence = fecha_inicio + timedelta(days=dias)
    else:
        r.plan_vence = fecha_inicio + timedelta(days=PLANES[plan_tipo]["dias"])
    db.session.commit()

    vence_str = r.plan_vence.strftime("%d/%m/%Y") if r.plan_vence else "—"
    flash(f"Plan '{r.plan}' asignado a {r.nombre}. Vence: {vence_str}", "success")
    return redirect(url_for("admin_panel"))


# ══════════════════════════════════════════════
#  PLANES (pública)
# ══════════════════════════════════════════════

@app.route("/planes")
def planes_page():
    return render_template("planes.html", planes=PLANES)


@app.route("/preview/plan-vencido")
@admin_required
def preview_plan_vencido():
    class MockRestaurante:
        nombre = "Mi Local Demo"
        whatsapp = "0991234567"
        pais = request.args.get("pais", "ecuador")
    return render_template("auth/plan_vencido.html",
                           restaurante=MockRestaurante(),
                           vencido_hace=3,
                           precios=PRECIOS.get(request.args.get("pais", "ecuador"), PRECIOS["ecuador"]),
                           pago_ecuador=PAGO_ECUADOR)


if __name__ == "__main__":
    app.run(debug=True, port=5001)
